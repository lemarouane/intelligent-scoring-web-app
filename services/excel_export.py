# ============================================================
# BOA INTELLIGENT CREDIT SCORING – EXCEL EXPORT
# ============================================================

import io
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


BLUE   = "003366"
GOLD   = "D4AF37"
GREEN  = "0E9F6E"
LIGHT  = "F5F7FA"
WHITE  = "FFFFFF"


def _border(style="thin"):
    s = Side(style=style, color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(
    categorie: str,
    final_score: float,
    risk_class: str,
    decision: str,
    prob_default: str,
    rows: list,
    ia_analysis: str = "",
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1 : Summary
    ws1 = wb.active
    ws1.title = "Résumé"

    ws1.merge_cells("A1:E1")
    c = ws1["A1"]
    c.value = "BANK OF AFRICA – Scoring de Crédit Professionnel"
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    meta = [
        ("Date", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Catégorie", categorie),
        ("Score Final", f"{final_score:.2f} / 100"),
        ("Classe de risque", f"Classe {risk_class}"),
        ("Décision", decision),
        ("Probabilité de défaut", prob_default),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws1[f"A{i}"] = k
        ws1[f"B{i}"] = v
        ws1[f"A{i}"].font = Font(bold=True, color=BLUE)
        ws1[f"A{i}"].fill = PatternFill("solid", fgColor=LIGHT)
        ws1[f"B{i}"].alignment = Alignment(horizontal="left")

    # ── Sheet 2 : Score Details
    ws2 = wb.create_sheet("Détail Scores")
    headers = ["Critère", "Poids (%)", "Score Partiel (/100)", "Score Pondéré", "Section"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for i, r in enumerate(rows, start=2):
        section = "BOA Commun" if r["poids"] <= 15 else "Spécifique"
        vals = [r["label"], r["poids"], r["partial"], r["weighted"], section]
        fill = PatternFill("solid", fgColor=LIGHT if i % 2 == 0 else WHITE)
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.fill = fill
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    # Total row
    total_row = len(rows) + 2
    ws2.cell(row=total_row, column=1, value="SCORE FINAL").font = Font(bold=True, color=WHITE)
    ws2.cell(row=total_row, column=4, value=final_score).font = Font(bold=True, color=WHITE)
    for col in range(1, 6):
        ws2.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor=GOLD)
        ws2.cell(row=total_row, column=col).border = _border()

    # Column widths
    col_widths = [40, 12, 22, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3 : IA Analysis
    if ia_analysis:
        ws3 = wb.create_sheet("Analyse IA")
        ws3.merge_cells("A1:B1")
        ws3["A1"] = "Analyse Intelligence Artificielle – Bank of Africa"
        ws3["A1"].font = Font(bold=True, size=12, color=WHITE)
        ws3["A1"].fill = PatternFill("solid", fgColor=BLUE)
        ws3["A1"].alignment = Alignment(horizontal="center")
        ws3.row_dimensions[1].height = 24

        for i, line in enumerate(ia_analysis.split("\n"), start=3):
            ws3.cell(row=i, column=1, value=line)
            ws3.row_dimensions[i].height = 15

        ws3.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()